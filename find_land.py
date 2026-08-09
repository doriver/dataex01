# 토지 실거래가(마스킹된 번지)에 매핑되는 토지특성정보(번지/PNU)를 찾아 xlsx로 저장
import csv
import sys
from pathlib import Path

import openpyxl

if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

XLSX_PATH = "data/raw/토지실거래가경기하남/토지(매매)_실거래가_20260809133253.xlsx"
CSV_PATH = "data/basis/토지특성정보경기csv/AL_D195_41_20260402.csv"
CSV_ENCODING = "cp949"
OUTPUT_PATH = "data/processed/토지_실거래가_매핑결과.xlsx"
SUCCESS_OUTPUT_PATH = "data/processed/토지_실거래가_매핑성공.xlsx"
AMBIGUOUS_OUTPUT_PATH = "data/processed/토지_실거래가_매핑후보다수.xlsx"
HEADER_ROW = 13
AREA_TOLERANCE = 0.02
MOUNTAIN_PREFIX = "산"
MOUNTAIN_JIBUN_CODE = "2"


def load_transactions(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[HEADER_ROW - 1])
    records = []
    for row in rows[HEADER_ROW:]:
        if row[0] is None:
            continue
        records.append(dict(zip(header, row)))
    return header, records


def load_land_index(path, sigungu_values):
    index = {}
    with open(path, encoding=CSV_ENCODING, newline="") as f:
        reader = csv.reader(f)
        fields = next(reader)
        for row in reader:
            record = dict(zip(fields, row))
            dong = record["법정동명"]
            if dong not in sigungu_values:
                continue
            index.setdefault(dong, []).append(record)
    return index


def parse_bunji(raw_bunji):
    raw_bunji = raw_bunji.strip()
    is_mountain = raw_bunji.startswith(MOUNTAIN_PREFIX)
    body = raw_bunji[len(MOUNTAIN_PREFIX):] if is_mountain else raw_bunji
    prefix = body.split("*", 1)[0]
    return is_mountain, prefix, len(body)


def find_candidates(record, land_index):
    """조건을 만족하는 토지특성정보 후보를 모두 찾아 반환한다 (0개, 1개, 여러 개 모두 가능)."""
    sigungu = record.get("시군구")
    candidates = land_index.get(sigungu)
    if not candidates:
        return []

    raw_bunji = record.get("번지")
    if not raw_bunji:
        return []
    is_mountain, prefix, total_len = parse_bunji(str(raw_bunji))

    jimok = record.get("지목")
    yongdo = record.get("용도지역")
    jibun_gubun = str(record.get("지분구분") or "").strip()

    try:
        contract_area = float(record.get("계약면적"))
    except (TypeError, ValueError):
        return []

    matches = []
    for land in candidates:
        jibun = land["지번"]
        if len(jibun) != total_len or not jibun.startswith(prefix):
            continue
        is_land_mountain = land["대장구분코드"] == MOUNTAIN_JIBUN_CODE
        if is_mountain != is_land_mountain:
            continue
        if land["지목명"] != jimok:
            continue
        if land["용도지역명1"] != yongdo:
            continue

        try:
            land_area = float(land["토지면적"])
        except (TypeError, ValueError):
            continue

        if jibun_gubun == "지분":
            if not (land_area > contract_area):
                continue
        else:
            if not (contract_area - AREA_TOLERANCE < land_area < contract_area + AREA_TOLERANCE):
                continue

        matches.append(land)

    return matches


def format_bunji(land, is_mountain):
    return MOUNTAIN_PREFIX + land["지번"] if is_mountain else land["지번"]


def find_match(record, land_index):
    matches = find_candidates(record, land_index)
    if len(matches) == 1:
        is_mountain, _, _ = parse_bunji(str(record.get("번지")).strip())
        land = matches[0]
        return {"번지": format_bunji(land, is_mountain), "pnu": land["고유번호"]}
    return None


CANDIDATE_COLUMNS = [
    ("후보_지번", "지번"),
    ("후보_대장구분명", "대장구분명"),
    ("후보_PNU", "고유번호"),
    ("후보_지목명", "지목명"),
    ("후보_용도지역명1", "용도지역명1"),
    ("후보_토지면적", "토지면적"),
]


def main():
    header, records = load_transactions(XLSX_PATH)
    print(f"실거래 데이터 {len(records)}건 로드")

    sigungu_values = {r.get("시군구") for r in records if r.get("시군구")}
    land_index = load_land_index(CSV_PATH, sigungu_values)
    print(f"토지특성정보 인덱싱 완료 (법정동 {len(land_index)}개, {sum(len(v) for v in land_index.values())}행)")

    success_records = []
    ambiguous_records = []

    for record in records:
        matches = find_candidates(record, land_index)
        if len(matches) == 1:
            is_mountain, _, _ = parse_bunji(str(record.get("번지")).strip())
            land = matches[0]
            record["번지(전체)"] = format_bunji(land, is_mountain)
            record["pnu"] = land["고유번호"]
            success_records.append(record)
        else:
            record["번지(전체)"] = None
            record["pnu"] = None
            if len(matches) >= 2:
                is_mountain, _, _ = parse_bunji(str(record.get("번지")).strip())
                ambiguous_records.append((record, matches, is_mountain))

    total_count = len(records)
    success_count = len(success_records)
    ambiguous_count = len(ambiguous_records)
    print(f"매핑 성공: {success_count} / {total_count}")
    print(f"후보 2개 이상(모호): {ambiguous_count} / {total_count}")

    out_header = header + ["번지(전체)", "pnu"]

    out_wb = openpyxl.Workbook()
    result_ws = out_wb.active
    result_ws.title = "매핑결과"
    result_ws.append(out_header)
    for record in records:
        result_ws.append([record.get(col) for col in out_header])

    summary_ws = out_wb.create_sheet("요약")
    summary_ws.append(["항목", "값"])
    summary_ws.append(["전체 건수", total_count])
    summary_ws.append(["매핑 성공 건수", success_count])
    summary_ws.append(["후보 2개 이상 건수", ambiguous_count])

    Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(OUTPUT_PATH)
    print(f"결과 저장: {OUTPUT_PATH}")

    success_wb = openpyxl.Workbook()
    success_ws = success_wb.active
    success_ws.title = "매핑성공"
    success_ws.append(out_header)
    for record in success_records:
        success_ws.append([record.get(col) for col in out_header])
    success_wb.save(SUCCESS_OUTPUT_PATH)
    print(f"매핑 성공 결과 저장: {SUCCESS_OUTPUT_PATH} ({success_count}건)")

    ambiguous_wb = openpyxl.Workbook()
    ambiguous_ws = ambiguous_wb.active
    ambiguous_ws.title = "후보다수"
    ambiguous_header = header + ["후보번호", "후보개수"] + [col for col, _ in CANDIDATE_COLUMNS] + ["후보_번지(전체)"]
    ambiguous_ws.append(ambiguous_header)
    for record, matches, is_mountain in ambiguous_records:
        base_values = [record.get(col) for col in header]
        for i, land in enumerate(matches, start=1):
            candidate_values = [land.get(src) for _, src in CANDIDATE_COLUMNS]
            ambiguous_ws.append(
                base_values + [i, len(matches)] + candidate_values + [format_bunji(land, is_mountain)]
            )
    ambiguous_wb.save(AMBIGUOUS_OUTPUT_PATH)
    print(f"후보 다수 결과 저장: {AMBIGUOUS_OUTPUT_PATH} ({ambiguous_count}건, {sum(len(m) for _, m, _ in ambiguous_records)}개 후보행)")


if __name__ == "__main__":
    main()
