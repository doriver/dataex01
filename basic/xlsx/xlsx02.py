# 토지 실거래가 데이터 중 '지분구분'이 '지분'인 행의 개수 출력
import sys

import openpyxl

# 콘솔 인코딩이 UTF-8이 아니면 한글 출력 깨짐 방지용으로 재설정
if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

XLSX_PATH = "data/raw/토지실거래가경기하남/토지(매매)_실거래가_20260809133253.xlsx"
HEADER_ROW = 13  # 1~12행은 안내문/검색조건, 13행이 실제 컬럼 헤더
TARGET_VALUE = "지분"  # 지분구분 컬럼에서 세고자 하는 값

# data_only=True: 수식이 아닌 계산된 값을 읽기 위함 (read_only=True는 이 파일에서 컬럼 수가 깨져서 사용하지 않음)
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[wb.sheetnames[0]]  # 첫 번째 시트
rows = list(ws.iter_rows(values_only=True))  # 셀 객체 대신 값만 튜플로 가져옴

header = list(rows[HEADER_ROW - 1])  # 13행이 실제 컬럼 헤더
gubun_idx = header.index("지분구분")  # '지분구분' 컬럼의 위치(인덱스) 찾기

# NO 값이 없는 행은 데이터가 아닌 빈 행이므로 제외
data_rows = [row for row in rows[HEADER_ROW:] if row[0] is not None]
# 지분구분 값이 '지분'인 행만 세어 개수 계산
jibun_count = sum(1 for row in data_rows if str(row[gubun_idx] or "").strip() == TARGET_VALUE)

print(f"전체 {len(data_rows)}건 중 지분구분='지분'인 행: {jibun_count}건")
