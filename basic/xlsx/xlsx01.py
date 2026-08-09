# 토지 실거래가 데이터 중 '해제사유발생일'이 있는(해제된) 행을 제외하고 저장
import sys
from pathlib import Path

import openpyxl

if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

XLSX_PATH = "data/raw/토지실거래가경기하남/토지(매매)_실거래가_20260809133253.xlsx"
OUTPUT_PATH = "data/processed/토지_실거래가_해제제외.xlsx"
HEADER_ROW = 13  # 1~12행은 안내문/검색조건, 13행이 실제 컬럼 헤더
EMPTY_VALUE = "-"  # 해제사유발생일이 없을 때 원본에 채워지는 값

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

header = list(rows[HEADER_ROW - 1])
cancel_date_idx = header.index("해제사유발생일")

# NO 값이 없는 행은 데이터가 아닌 빈 행이므로 제외
data_rows = [row for row in rows[HEADER_ROW:] if row[0] is not None]
# 해제사유발생일이 비어있거나 '-'인(=해제되지 않은) 행만 남긴다
kept_rows = [row for row in data_rows if str(row[cancel_date_idx] or "").strip() in ("", EMPTY_VALUE)]

print(f"전체 {len(data_rows)}건 중 해제 {len(data_rows) - len(kept_rows)}건 제외, {len(kept_rows)}건 저장")

out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "실거래가"
out_ws.append(header)
for row in kept_rows:
    out_ws.append(row)

Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)
out_wb.save(OUTPUT_PATH)
print(f"결과 저장: {OUTPUT_PATH}")
